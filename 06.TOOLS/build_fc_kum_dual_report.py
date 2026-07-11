from __future__ import annotations

import csv
import json
import re
import sqlite3
import traceback
import warnings
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
warnings.filterwarnings("ignore", message=".*Conditional Formatting extension.*")

ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "01.DB" / "close_report.sqlite"
SOURCE_XLSX = ROOT / "03.DATA" / "raw" / "current" / "main.xlsx"
OUTPUT_XLSX = ROOT / "04.REPORTS" / "outputs" / "CloseReport_FC_KUM_DB_REPORTS.xlsx"
DOCS_DIR = ROOT / "05.DOCS"

VALIDATION_PATH = DOCS_DIR / "FC_KUM_VALIDATION.CloseReport.csv"
REVIEW_PATH = DOCS_DIR / "FC_KUM_REVIEW_REQUIRED.CloseReport.csv"
CONTRACT_PATH = DOCS_DIR / "FC_KUM_REPORT_CONTRACT.CloseReport.csv"

TARGET_SHEET = "F.CKUM"
DETAIL_SHEET = "F.D.KUM.01"
DETAIL_DB_SHEET = "F.D.KUM.01.DB"
REQUIRED_SHEETS = [
    "CONTROL",
    "F.D.KUM.01.DB",
    "F.CKUM.EFFICIENT",
    "F.CKUM.OS",
    "VALIDATION.RESULTS",
    "REVIEW.REQUIRED",
    "REPORT.CONTRACT",
]

REVIEW_HEADERS = [
    "file",
    "sheet",
    "cell",
    "formula_or_value",
    "dependency_type",
    "reason",
    "classification",
    "display_value",
]
VALIDATION_HEADERS = ["scope", "item", "expected", "actual", "status", "notes"]
CONTRACT_HEADERS = ["contract_key", "contract_value"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_DB = PatternFill("solid", fgColor="C6EFCE")
FILL_REVIEW = PatternFill("solid", fgColor="FFC7CE")
FILL_OUTPUT_ONLY = PatternFill("solid", fgColor="E7E6E6")
FILL_LOCAL = PatternFill("solid", fgColor="D9EAD3")

BLOCKED_SHEET_REFS = {"F.CMC", "F.C.KMT", "F.CKMT", "F.C.KUM"}
SHEET_REF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_. ]+))!")


def safe_text(value: object) -> str:
    return "" if value is None else str(value)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def style_header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for index in range(1, len(headers) + 1):
        cell = ws.cell(1, index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = CELL_BORDER


def copy_cell_style(src_cell, dst_cell) -> None:
    if isinstance(src_cell, MergedCell):
        return
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)


def copy_dimensions(src_ws, dst_ws) -> None:
    for column_index in range(1, src_ws.max_column + 1):
        letter = get_column_letter(column_index)
        if letter in src_ws.column_dimensions:
            src_dim = src_ws.column_dimensions[letter]
            dst_dim = dst_ws.column_dimensions[letter]
            dst_dim.width = src_dim.width
            dst_dim.hidden = src_dim.hidden
    for row_index in range(1, src_ws.max_row + 1):
        if row_index in src_ws.row_dimensions:
            src_dim = src_ws.row_dimensions[row_index]
            dst_dim = dst_ws.row_dimensions[row_index]
            dst_dim.height = src_dim.height
            dst_dim.hidden = src_dim.hidden


def copy_sheet_template(style_ws, value_ws, dst_ws) -> None:
    copy_dimensions(style_ws, dst_ws)
    dst_ws.sheet_view.showGridLines = style_ws.sheet_view.showGridLines
    dst_ws.freeze_panes = style_ws.freeze_panes
    for merged_range in style_ws.merged_cells.ranges:
        try:
            dst_ws.merge_cells(str(merged_range))
        except Exception:
            continue
    for row in range(1, style_ws.max_row + 1):
        for col in range(1, style_ws.max_column + 1):
            src_cell = style_ws.cell(row, col)
            if isinstance(src_cell, MergedCell):
                continue
            dst_cell = dst_ws.cell(row, col)
            dst_cell.value = value_ws.cell(row, col).value
            copy_cell_style(src_cell, dst_cell)


def extract_sheet_names(formula: str) -> list[str]:
    sheet_names: list[str] = []
    for match in SHEET_REF_RE.finditer(formula):
        name = (match.group(1) or match.group(2) or "").strip()
        if name and name not in sheet_names:
            sheet_names.append(name)
    return sheet_names


def classify_formula(formula: str) -> tuple[str, str, str]:
    upper = formula.upper()
    if "#REF!" in upper:
        return "REF_ERROR", "REF_ERROR", "Formula contains #REF!."
    if "[" in formula and "]" in formula:
        return "EXTERNAL_REFERENCE", "EXTERNAL", "Formula contains external workbook link."

    sheet_names = set(extract_sheet_names(formula))
    if DETAIL_SHEET in sheet_names:
        return "DETAIL_DEPENDENCY", "DETAIL_DB", "Formula references F.D.KUM.01 and can be retargeted safely."
    if "BASES" in upper and "SUMIFS" in upper:
        return "BASES_SUMIFS_PENDING_SQL", "BASES_DB_PENDING", "SUMIFS against Bases is not reconstructed to SQL in this pass."
    if any(name in BLOCKED_SHEET_REFS for name in sheet_names):
        return "CROSS_REPORT_DEPENDENCY", "CROSS_REPORT", "Formula references a blocked report sheet."
    if "!" in formula:
        return "FORMULA_OTHER_SHEET", "OTHER_SHEET", "Formula references another non-local sheet."
    return "LOCAL_FORMULA", "LOCAL", "Formula is local to F.CKUM."


def classify_cell(cell) -> tuple[str, str, str]:
    value = cell.value
    if value is None:
        return "Output_Only", "VISUAL_LAYOUT", "Blank visual/output-only cell."
    if isinstance(value, str) and value.startswith("="):
        return classify_formula(value)
    return "Output_Only", "VALUE", "Static text/date/number output-only cell."


def retarget_detail_formula(formula: str) -> str:
    output = formula.replace("'F.D.KUM.01'!", "'F.D.KUM.01.DB'!")
    output = output.replace("F.D.KUM.01!", "'F.D.KUM.01.DB'!")
    return output


def validate_db(validation_rows: list[dict[str, str]]) -> int:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"Required database missing: {DB_PATH}")

    connection = sqlite3.connect(DB_PATH)
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='movements'")
        has_movements = cursor.fetchone() is not None
        validation_rows.append(
            {
                "scope": "DB",
                "item": "movements_table",
                "expected": "exists",
                "actual": "exists" if has_movements else "missing",
                "status": "PASS" if has_movements else "FAIL",
                "notes": "",
            }
        )
        if not has_movements:
            raise RuntimeError("Required table movements was not found.")

        cursor.execute("SELECT COUNT(*) FROM movements")
        row_count = int(cursor.fetchone()[0])
        validation_rows.append(
            {
                "scope": "DB",
                "item": "movements_row_count",
                "expected": "4501",
                "actual": str(row_count),
                "status": "PASS" if row_count == 4501 else "WARN",
                "notes": "",
            }
        )
        return row_count
    finally:
        connection.close()


def build_contract_rows(run_id: str) -> list[dict[str, str]]:
    return [
        {"contract_key": "report_id", "contract_value": "FC_KUM_P0_PREBUILD"},
        {"contract_key": "run_id", "contract_value": run_id},
        {"contract_key": "target_report", "contract_value": "PDF page 4 / F.CKUM / Kumquat Holding"},
        {"contract_key": "status", "contract_value": "RUNNABLE_PRECANONICAL_PASS_WITH_REVIEW"},
        {"contract_key": "database", "contract_value": str(DB_PATH)},
        {"contract_key": "source_workbook", "contract_value": str(SOURCE_XLSX)},
        {"contract_key": "compare_sheet", "contract_value": TARGET_SHEET},
        {"contract_key": "compare_detail_sheet", "contract_value": DETAIL_SHEET},
        {"contract_key": "output_workbook", "contract_value": str(OUTPUT_XLSX)},
        {"contract_key": "fcmc_policy", "contract_value": "F.CMC remains blocked until KMT and KUM are stable."},
    ]


def add_control_sheet(ws, run_id: str, movement_rows: int, review_rows: int) -> None:
    style_header_row(ws, ["field", "value"])
    rows = [
        ("run_id", run_id),
        ("generated_at", datetime.now().isoformat(timespec="seconds")),
        ("target_report", TARGET_SHEET),
        ("source_detail_sheet", DETAIL_SHEET),
        ("source_workbook", str(SOURCE_XLSX)),
        ("database", str(DB_PATH)),
        ("output_workbook", str(OUTPUT_XLSX)),
        ("movements_rows", str(movement_rows)),
        ("review_required_rows_initial", str(review_rows)),
        ("status_policy", "PASS_WITH_REVIEW is acceptable until cross-report dependencies are resolved."),
    ]
    for key, value in rows:
        ws.append([key, value])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 130


def add_validation_sheet(ws, validation_rows: list[dict[str, str]]) -> None:
    style_header_row(ws, VALIDATION_HEADERS)
    for row in validation_rows:
        ws.append([row.get(header, "") for header in VALIDATION_HEADERS])
    ws.freeze_panes = "A2"


def add_review_sheet(ws, review_rows: list[dict[str, str]]) -> None:
    style_header_row(ws, REVIEW_HEADERS)
    for row in review_rows:
        ws.append([row.get(header, "") for header in REVIEW_HEADERS])
    ws.freeze_panes = "A2"


def add_contract_sheet(ws, contract_rows: list[dict[str, str]]) -> None:
    style_header_row(ws, CONTRACT_HEADERS)
    for row in contract_rows:
        ws.append([row["contract_key"], row["contract_value"]])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 130


def populate_visual_sheets(
    source_formula_ws,
    source_value_ws,
    os_ws,
    eff_ws,
    review_rows: list[dict[str, str]],
) -> None:
    copy_sheet_template(source_formula_ws, source_value_ws, os_ws)
    copy_sheet_template(source_formula_ws, source_value_ws, eff_ws)

    for row in range(1, source_formula_ws.max_row + 1):
        for col in range(1, source_formula_ws.max_column + 1):
            src_cell = source_formula_ws.cell(row, col)
            if isinstance(src_cell, MergedCell):
                continue

            coord = f"{get_column_letter(col)}{row}"
            display_value = source_value_ws.cell(row, col).value
            source_value = src_cell.value
            classification, dependency_type, reason = classify_cell(src_cell)

            os_cell = os_ws.cell(row, col)
            eff_cell = eff_ws.cell(row, col)

            if classification == "DETAIL_DEPENDENCY":
                formula = retarget_detail_formula(str(source_value))
                os_cell.value = formula
                eff_cell.value = formula
                os_cell.fill = FILL_DB
                eff_cell.fill = copy(FILL_DB)
            elif classification == "LOCAL_FORMULA":
                os_cell.value = source_value
                eff_cell.value = source_value
                os_cell.fill = FILL_LOCAL
                eff_cell.fill = copy(FILL_LOCAL)
            elif classification == "Output_Only":
                os_cell.value = display_value
                eff_cell.value = display_value
                os_cell.fill = FILL_OUTPUT_ONLY
                eff_cell.fill = copy(FILL_OUTPUT_ONLY)
            else:
                os_cell.value = display_value
                eff_cell.value = display_value
                os_cell.fill = FILL_REVIEW
                eff_cell.fill = copy(FILL_REVIEW)
                review_rows.append(
                    {
                        "file": str(SOURCE_XLSX),
                        "sheet": TARGET_SHEET,
                        "cell": coord,
                        "formula_or_value": safe_text(source_value),
                        "dependency_type": dependency_type,
                        "reason": reason,
                        "classification": classification,
                        "display_value": safe_text(display_value),
                    }
                )


def verify_output_workbook(validation_rows: list[dict[str, str]]) -> None:
    workbook = load_workbook(OUTPUT_XLSX, read_only=True, data_only=False)
    sheet_names = workbook.sheetnames
    validation_rows.append(
        {
            "scope": "WORKBOOK",
            "item": "reopen_after_save",
            "expected": ", ".join(REQUIRED_SHEETS),
            "actual": ", ".join(sheet_names),
            "status": "PASS" if sheet_names == REQUIRED_SHEETS else "FAIL",
            "notes": "Workbook reopened after save.",
        }
    )
    workbook.close()


def build() -> dict[str, object]:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Required source workbook missing: {SOURCE_XLSX}")

    run_id = "FC_KUM_DUAL_REPORT_" + datetime.now().strftime("%Y%m%d_%H%M%S")
    validation_rows: list[dict[str, str]] = []
    review_rows: list[dict[str, str]] = []

    movement_rows = validate_db(validation_rows)
    source_formula_wb = load_workbook(SOURCE_XLSX, data_only=False, keep_links=True)
    source_value_wb = load_workbook(SOURCE_XLSX, data_only=True, keep_links=True)

    for required in (TARGET_SHEET, DETAIL_SHEET):
        exists = required in source_formula_wb.sheetnames
        validation_rows.append(
            {
                "scope": "SOURCE",
                "item": f"sheet_{required}",
                "expected": "exists",
                "actual": "exists" if exists else "missing",
                "status": "PASS" if exists else "FAIL",
                "notes": str(SOURCE_XLSX),
            }
        )
        if not exists:
            raise RuntimeError(f"Required sheet missing in source workbook: {required}")

    workbook = Workbook()
    workbook.remove(workbook.active)
    control_ws = workbook.create_sheet("CONTROL")
    detail_ws = workbook.create_sheet("F.D.KUM.01.DB")
    efficient_ws = workbook.create_sheet("F.CKUM.EFFICIENT")
    os_ws = workbook.create_sheet("F.CKUM.OS")
    validation_ws = workbook.create_sheet("VALIDATION.RESULTS")
    review_ws = workbook.create_sheet("REVIEW.REQUIRED")
    contract_ws = workbook.create_sheet("REPORT.CONTRACT")

    copy_sheet_template(source_formula_wb[DETAIL_SHEET], source_value_wb[DETAIL_SHEET], detail_ws)
    populate_visual_sheets(
        source_formula_wb[TARGET_SHEET],
        source_value_wb[TARGET_SHEET],
        os_ws,
        efficient_ws,
        review_rows,
    )

    validation_rows.append(
        {
            "scope": "WORKBOOK",
            "item": "required_sheets",
            "expected": ", ".join(REQUIRED_SHEETS),
            "actual": ", ".join(workbook.sheetnames),
            "status": "PASS" if workbook.sheetnames == REQUIRED_SHEETS else "FAIL",
            "notes": "",
        }
    )
    status = "PASS_WITH_REVIEW" if review_rows else "PASS"
    validation_rows.append(
        {
            "scope": "WORKBOOK",
            "item": "review_required_rows",
            "expected": "0 for clean pass",
            "actual": str(len(review_rows)),
            "status": "WARN" if review_rows else "PASS",
            "notes": "PASS_WITH_REVIEW is acceptable while unresolved dependencies remain registered.",
        }
    )
    validation_rows.append(
        {
            "scope": "WORKBOOK",
            "item": "final_status",
            "expected": "PASS or PASS_WITH_REVIEW",
            "actual": status,
            "status": "PASS",
            "notes": "F.CKUM is runnable precanonical and not yet closed.",
        }
    )

    add_control_sheet(control_ws, run_id, movement_rows, len(review_rows))
    add_validation_sheet(validation_ws, validation_rows)
    add_review_sheet(review_ws, review_rows)
    contract_rows = build_contract_rows(run_id)
    add_contract_sheet(contract_ws, contract_rows)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)
    verify_output_workbook(validation_rows)

    write_csv(VALIDATION_PATH, VALIDATION_HEADERS, validation_rows)
    write_csv(REVIEW_PATH, REVIEW_HEADERS, review_rows)
    write_csv(CONTRACT_PATH, CONTRACT_HEADERS, contract_rows)

    source_formula_wb.close()
    source_value_wb.close()

    return {
        "status": status,
        "review_rows": len(review_rows),
        "output_workbook": str(OUTPUT_XLSX),
        "validation_path": str(VALIDATION_PATH),
        "review_path": str(REVIEW_PATH),
        "contract_path": str(CONTRACT_PATH),
    }


def main() -> int:
    try:
        result = build()
        print(json.dumps(result, indent=2))
        return 0
    except Exception:
        print(traceback.format_exc())
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
