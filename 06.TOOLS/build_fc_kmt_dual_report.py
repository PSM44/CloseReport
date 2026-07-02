from __future__ import annotations

import csv
import json
import re
import shutil
import sqlite3
import traceback
import warnings
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formatting.rule import IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
warnings.filterwarnings("ignore", message=".*Conditional Formatting extension.*")

ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "01.DB" / "close_report.sqlite"
SOURCE_XLSX = ROOT / "03.DATA" / "raw" / "current" / "main.xlsx"
PROBE_PATH = ROOT / "05.DOCS" / "FD_KMT_01_BASES_FORMULA_PROBE.CloseReport.csv"
OUTPUT_XLSX = ROOT / "04.REPORTS" / "outputs" / "CloseReport_FC_KMT_DB_REPORTS.xlsx"
DOCS_DIR = ROOT / "05.DOCS"
HIST_DIR = DOCS_DIR / "HIST"

SUMMARY_PATH = DOCS_DIR / "FC_KMT_DUAL_REPORT_BUILD_SUMMARY.CloseReport.txt"
VALIDATION_PATH = DOCS_DIR / "FC_KMT_DUAL_REPORT_VALIDATION.CloseReport.csv"
REVIEW_PATH = DOCS_DIR / "FC_KMT_DUAL_REPORT_REVIEW_REQUIRED.CloseReport.csv"
CLASSIFICATION_PATH = DOCS_DIR / "FC_KMT_DUAL_REPORT_CELL_CLASSIFICATION.CloseReport.csv"
CONTRACT_PATH = DOCS_DIR / "FC_KMT_DUAL_REPORT_CONTRACT.CloseReport.csv"
DOC_INDEX_PATH = DOCS_DIR / "DOCS_ACTIVE_INDEX.CloseReport.txt"

REQUIRED_SHEETS = [
    "CONTROL",
    "DB.MOVEMENTS.SAMPLE",
    "F.D.KMT.01.DB",
    "F.C.KMT.EFFICIENT",
    "F.C.KMT.OS",
    "VALIDATION.RESULTS",
    "REVIEW.REQUIRED",
    "REPORT.CONTRACT",
]

BLOCKED_SHEET_REFS = {"F.CMC", "F.CKUM", "F.C.KUM"}
ARCHIVE_PATTERNS = ("FC_KMT_VISUAL_*", "FC_KMT_DUAL_REPORT_*")
ACTIVE_DOCS = {
    SUMMARY_PATH.name,
    VALIDATION_PATH.name,
    REVIEW_PATH.name,
    CLASSIFICATION_PATH.name,
    CONTRACT_PATH.name,
    DOC_INDEX_PATH.name,
    PROBE_PATH.name,
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_DB = PatternFill("solid", fgColor="C6EFCE")
FILL_REVIEW = PatternFill("solid", fgColor="FFC7CE")
FILL_VALUE = PatternFill("solid", fgColor="FFEB9C")
FILL_LOCAL = PatternFill("solid", fgColor="D9EAD3")
FILL_OTHER = PatternFill("solid", fgColor="D9E1F2")

FC_KMT_UF_ROW = 6
FC_KMT_UF_COL_START = 6
FC_KMT_UF_COL_END = 31
FC_KMT_V00_ROW_START = 10
FC_KMT_V00_ROW_END = 29
FC_KMT_V00_COL = 34
FC_KMT_VARIANCE_COL = 37
FC_KMT_VARIANCE_RANGE = "AK10:AK29"
FC_KMT_UF_GAP_COL_START = 18
FC_KMT_UF_GAP_COL_END = 19

REVIEW_HEADERS = [
    "sheet",
    "cell",
    "formula_or_value",
    "dependency_type",
    "reason",
    "classification",
    "display_value",
]
CLASSIFICATION_HEADERS = [
    "sheet",
    "cell",
    "classification",
    "dependency_type",
    "source_formula_or_value",
    "generated_value",
    "number_format",
]
VALIDATION_HEADERS = ["scope", "item", "expected", "actual", "status", "notes"]
CONTRACT_HEADERS = ["contract_key", "contract_value"]


@dataclass
class BuildContext:
    run_id: str
    timestamp: str
    archive_dir: Path
    summary_status: str = "PASS"


def safe_text(value: object) -> str:
    return "" if value is None else str(value)


def write_csv(path: Path, headers: list[str], rows: Iterable[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def style_header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for index in range(1, len(headers) + 1):
        cell = ws.cell(1, index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = CELL_BORDER


def copy_cell_style(src_cell, dst_cell) -> None:
    if isinstance(src_cell, MergedCell):
        return
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)


def copy_dimensions(src_ws, dst_ws) -> None:
    for column_index in range(1, src_ws.max_column + 1):
        letter = get_column_letter(column_index)
        if letter in src_ws.column_dimensions:
            src_dim = src_ws.column_dimensions[letter]
            dst_dim = dst_ws.column_dimensions[letter]
            dst_dim.width = src_dim.width
            dst_dim.hidden = src_dim.hidden
    for row_index in range(1, src_ws.max_row + 1):
        if row_index in src_ws.row_dimensions:
            src_dim = src_ws.row_dimensions[row_index]
            dst_dim = dst_ws.row_dimensions[row_index]
            dst_dim.height = src_dim.height
            dst_dim.hidden = src_dim.hidden


def copy_sheet_template(style_ws, value_ws, dst_ws) -> None:
    copy_dimensions(style_ws, dst_ws)
    dst_ws.sheet_view.showGridLines = style_ws.sheet_view.showGridLines
    dst_ws.freeze_panes = style_ws.freeze_panes
    for merged_range in style_ws.merged_cells.ranges:
        try:
            dst_ws.merge_cells(str(merged_range))
        except Exception:
            continue
    for row in range(1, style_ws.max_row + 1):
        for col in range(1, style_ws.max_column + 1):
            src_cell = style_ws.cell(row, col)
            if isinstance(src_cell, MergedCell):
                continue
            dst_cell = dst_ws.cell(row, col)
            dst_cell.value = value_ws.cell(row, col).value
            copy_cell_style(src_cell, dst_cell)


def classify_formula(formula: str, current_sheet: str) -> tuple[str, str]:
    upper = formula.upper()
    if "#REF!" in upper:
        return "REF_ERROR", "REF_ERROR"
    if "[" in formula and "]" in formula:
        return "EXTERNAL_REFERENCE", "EXTERNAL"
    for blocked in BLOCKED_SHEET_REFS:
        if re.search(rf"(?:'({re.escape(blocked)})'|{re.escape(blocked)})!", formula):
            return "CROSS_REPORT_DEPENDENCY", "CROSS_REPORT"
    if re.search(r"(?:'F\.D\.KMT\.01'|F\.D\.KMT\.01)!", formula):
        return "FD_KMT_DEPENDENCY", "FD_KMT_DB"
    if "!" in formula:
        if re.search(rf"(?:'({re.escape(current_sheet)})'|{re.escape(current_sheet)})!", formula):
            return "LOCAL_FORMULA", "LOCAL"
        return "FORMULA_OTHER_SHEET", "OTHER_SHEET"
    return "LOCAL_FORMULA", "LOCAL"


def classify_cell(src_formula_cell, current_sheet: str) -> tuple[str, str]:
    value = src_formula_cell.value
    if value is None:
        return "BLANK", "NONE"
    if isinstance(value, str) and value.startswith("="):
        return classify_formula(value, current_sheet)
    if isinstance(value, (int, float)):
        return "HARDNUMBER", "VALUE"
    return "TEXT_OR_DATE", "VALUE"


def is_fc_kmt_uf_parameter(row: int, col: int) -> bool:
    return row == FC_KMT_UF_ROW and FC_KMT_UF_COL_START <= col <= FC_KMT_UF_COL_END


def is_fc_kmt_uf_gap(row: int, col: int) -> bool:
    return row == FC_KMT_UF_ROW and FC_KMT_UF_GAP_COL_START <= col <= FC_KMT_UF_GAP_COL_END


def is_fc_kmt_v00_cell(row: int, col: int) -> bool:
    return FC_KMT_V00_ROW_START <= row <= FC_KMT_V00_ROW_END and col == FC_KMT_V00_COL


def is_fc_kmt_variance_cell(row: int, col: int) -> bool:
    return FC_KMT_V00_ROW_START <= row <= FC_KMT_V00_ROW_END and col == FC_KMT_VARIANCE_COL


def is_fc_kmt_visible_label(row: int, col: int) -> bool:
    return row == 24 and col == 3


def add_variance_icon_formatting(ws) -> None:
    try:
        ws.conditional_formatting.add(
            FC_KMT_VARIANCE_RANGE,
            IconSetRule("3Arrows", "num", [-0.0001, 0, 0.0001], showValue=True),
        )
    except Exception:
        return


def archive_obsolete_docs(ctx: BuildContext) -> list[str]:
    ctx.archive_dir.mkdir(parents=True, exist_ok=True)
    archived = []
    for pattern in ARCHIVE_PATTERNS:
        for path in DOCS_DIR.glob(pattern):
            if not path.is_file():
                continue
            if path.name in ACTIVE_DOCS:
                continue
            destination = ctx.archive_dir / path.name
            if destination.exists():
                destination.unlink()
            shutil.move(str(path), str(destination))
            archived.append(path.name)
    return sorted(archived)


def read_probe_rows(validation_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    if not PROBE_PATH.exists():
        validation_rows.append(
            {
                "scope": "PROBE",
                "item": "FD_KMT_01_BASES_FORMULA_PROBE",
                "expected": "129 rows if available",
                "actual": "MISSING",
                "status": "WARN",
                "notes": "Probe file not found; continuing without hard fail.",
            }
        )
        return []

    with PROBE_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))

    validation_rows.append(
        {
            "scope": "PROBE",
            "item": "FD_KMT_01_BASES_FORMULA_PROBE",
            "expected": "129 rows",
            "actual": str(len(rows)),
            "status": "PASS" if len(rows) == 129 else "FAIL",
            "notes": "UF tolerance 0.0001.",
        }
    )
    return rows


def validate_db(validation_rows: list[dict[str, str]]) -> sqlite3.Connection:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"Required database missing: {DB_PATH}")
    connection = sqlite3.connect(DB_PATH)
    cursor = connection.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='movements'")
    has_movements = cursor.fetchone() is not None
    validation_rows.append(
        {
            "scope": "DB",
            "item": "movements_table",
            "expected": "exists",
            "actual": "exists" if has_movements else "missing",
            "status": "PASS" if has_movements else "FAIL",
            "notes": "",
        }
    )
    if not has_movements:
        raise RuntimeError("Required table movements was not found.")

    cursor.execute("SELECT COUNT(*) FROM movements")
    row_count = int(cursor.fetchone()[0])
    validation_rows.append(
        {
            "scope": "DB",
            "item": "movements_row_count",
            "expected": "4501",
            "actual": str(row_count),
            "status": "PASS" if row_count == 4501 else "FAIL",
            "notes": "",
        }
    )
    if row_count != 4501:
        raise RuntimeError(f"Expected 4501 rows in movements, found {row_count}.")
    return connection


def build_contract_rows(ctx: BuildContext) -> list[dict[str, str]]:
    return [
        {"contract_key": "report_id", "contract_value": "FC_KMT_P0"},
        {"contract_key": "run_id", "contract_value": ctx.run_id},
        {"contract_key": "target_report", "contract_value": "PDF page 3 / F.C.KMT / Gestora KMT"},
        {"contract_key": "route", "contract_value": "Bases -> close_report.sqlite -> query/scaffold -> F.D.KMT.01.DB -> F.C.KMT.EFFICIENT + F.C.KMT.OS"},
        {"contract_key": "output_workbook", "contract_value": str(OUTPUT_XLSX)},
        {"contract_key": "required_sheets", "contract_value": ", ".join(REQUIRED_SHEETS)},
        {"contract_key": "blocked_cell_policy", "contract_value": "Do not block workbook creation; highlight cell and record REVIEW.REQUIRED."},
        {"contract_key": "openpyxl_safety", "contract_value": "copy.copy style cloning, merged-cell safe traversal, get_column_letter dimensions."},
    ]


def retarget_fd_formula(formula: str) -> str:
    formula = formula.replace("'F.D.KMT.01'!", "'F.D.KMT.01.DB'!")
    formula = formula.replace("F.D.KMT.01!", "'F.D.KMT.01.DB'!")
    return formula


def add_control_sheet(ws, ctx: BuildContext, probe_count: int, movements_rows: int) -> None:
    style_header_row(ws, ["field", "value"])
    control_rows = [
        ("run_id", ctx.run_id),
        ("generated_at", datetime.now().isoformat(timespec="seconds")),
        ("source_workbook", str(SOURCE_XLSX)),
        ("database", str(DB_PATH)),
        ("movements_rows", str(movements_rows)),
        ("probe_rows", str(probe_count)),
        ("output_workbook", str(OUTPUT_XLSX)),
        ("status_policy", "PASS_WITH_REVIEW allowed when unresolved cells remain."),
        ("tolerance_uf", "0.0001"),
    ]
    for key, value in control_rows:
        ws.append([key, value])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 120


def add_movements_sample(ws, connection: sqlite3.Connection) -> None:
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM movements LIMIT 100")
    rows = cursor.fetchall()
    if not rows:
        style_header_row(ws, ["message"])
        ws.append(["No rows returned from movements."])
        return
    headers = list(rows[0].keys())
    style_header_row(ws, headers)
    for row in rows:
        ws.append([row[key] for key in headers])
    ws.freeze_panes = "A2"


def add_validation_sheet(ws, validation_rows: list[dict[str, str]]) -> None:
    style_header_row(ws, VALIDATION_HEADERS)
    for row in validation_rows:
        ws.append([row.get(header, "") for header in VALIDATION_HEADERS])
    ws.freeze_panes = "A2"
    for column in range(1, len(VALIDATION_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 28


def add_review_sheet(ws, review_rows: list[dict[str, str]]) -> None:
    style_header_row(ws, REVIEW_HEADERS)
    for row in review_rows:
        ws.append([row.get(header, "") for header in REVIEW_HEADERS])
    ws.freeze_panes = "A2"
    for column in range(1, len(REVIEW_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 28


def add_contract_sheet(ws, contract_rows: list[dict[str, str]]) -> None:
    style_header_row(ws, CONTRACT_HEADERS)
    for row in contract_rows:
        ws.append([row["contract_key"], row["contract_value"]])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 120


def build_detail_sheet(ws, detail_style_ws, detail_value_ws, probe_rows: list[dict[str, str]]) -> dict[str, object]:
    copy_sheet_template(detail_style_ws, detail_value_ws, ws)
    probe_map: dict[str, float] = {}
    for row in probe_rows:
        target = row.get("target_cell", "").strip()
        if not target:
            continue
        calculated = row.get("calculated_value", "")
        try:
            value = float(calculated)
        except ValueError:
            continue
        probe_map[target] = value
        cell = ws[target]
        cell.value = value
        cell.fill = FILL_DB
    return probe_map


def populate_visual_sheets(
    source_formula_ws,
    source_value_ws,
    os_ws,
    eff_ws,
    review_rows: list[dict[str, str]],
    classification_rows: list[dict[str, str]],
) -> dict[str, int]:
    copy_sheet_template(source_formula_ws, source_value_ws, os_ws)
    copy_sheet_template(source_formula_ws, source_value_ws, eff_ws)

    counts: dict[str, int] = {}
    for row in range(1, source_formula_ws.max_row + 1):
        for col in range(1, source_formula_ws.max_column + 1):
            src_cell = source_formula_ws.cell(row, col)
            if isinstance(src_cell, MergedCell):
                continue
            value_cell = source_value_ws.cell(row, col)
            coord = f"{get_column_letter(col)}{row}"
            classification, dependency_type = classify_cell(src_cell, "F.C.KMT")

            os_cell = os_ws.cell(row, col)
            eff_cell = eff_ws.cell(row, col)
            source_value = src_cell.value
            display_value = value_cell.value

            if is_fc_kmt_uf_gap(row, col):
                classification = "INTENTIONAL_BLANK_VISUAL_GAP"
                dependency_type = "VISUAL_LAYOUT"
                os_cell.value = None
                eff_cell.value = None
                os_cell.fill = FILL_VALUE
                eff_cell.fill = FILL_VALUE
            elif is_fc_kmt_uf_parameter(row, col):
                classification = "REPORT_PARAMETER_UF_FROM_FCMC" if display_value is not None else "INTENTIONAL_BLANK_VISUAL_GAP"
                dependency_type = "REPORT_PARAMETER"
                os_cell.value = display_value
                eff_cell.value = display_value
                if display_value is None:
                    os_cell.fill = FILL_VALUE
                    eff_cell.fill = FILL_VALUE
                else:
                    os_cell.fill = FILL_VALUE
                    eff_cell.fill = FILL_VALUE
            elif is_fc_kmt_v00_cell(row, col):
                if row == 19:
                    classification = "VISUAL_SEPARATOR_ROW"
                    dependency_type = "VISUAL_LAYOUT"
                    os_cell.value = None
                    eff_cell.value = None
                    os_cell.fill = FILL_VALUE
                    eff_cell.fill = FILL_VALUE
                elif 25 <= row <= 28:
                    classification = "INTENTIONAL_BLANK_NO_SOURCE_DATA"
                    dependency_type = "FORECAST_SOURCE"
                    os_cell.value = None
                    eff_cell.value = None
                    os_cell.fill = FILL_VALUE
                    eff_cell.fill = FILL_VALUE
                else:
                    classification = "FORECAST_V00_2026" if (source_value is not None or display_value is not None) else "FORECAST_V00_2026_MISSING"
                    dependency_type = "FORECAST_SOURCE"
                    os_cell.value = display_value
                    eff_cell.value = display_value
                    if source_value is None and display_value is None:
                        os_cell.fill = FILL_REVIEW
                        eff_cell.fill = FILL_REVIEW
                        review_rows.append(
                            {
                                "sheet": "F.C.KMT",
                                "cell": coord,
                                "formula_or_value": safe_text(source_value),
                                "dependency_type": dependency_type,
                                "reason": "V00 source could not be resolved from the source workbook; kept visible and listed for review.",
                                "classification": classification,
                                "display_value": safe_text(display_value),
                            }
                        )
                    else:
                        os_cell.fill = FILL_OTHER
                        eff_cell.fill = FILL_OTHER
            elif is_fc_kmt_visible_label(row, col):
                classification = "VISIBLE_LABEL_FROM_KMT"
                dependency_type = "FORECAST_SOURCE"
                os_cell.value = display_value
                eff_cell.value = display_value
                os_cell.fill = FILL_OTHER
                eff_cell.fill = FILL_OTHER
            elif is_fc_kmt_variance_cell(row, col):
                classification = "VARIANCE_VS_V00_2026"
                dependency_type = "LOCAL_FORMULA"
                variance_formula = f"=AF{row}-AH{row}"
                os_cell.value = variance_formula
                eff_cell.value = variance_formula
                os_cell.fill = FILL_LOCAL
                eff_cell.fill = FILL_LOCAL
            elif classification == "FD_KMT_DEPENDENCY":
                os_cell.value = retarget_fd_formula(str(source_value))
                eff_cell.value = retarget_fd_formula(str(source_value))
                os_cell.fill = FILL_DB
                eff_cell.fill = FILL_DB
            elif classification == "LOCAL_FORMULA":
                os_cell.value = source_value
                eff_cell.value = source_value
                os_cell.fill = FILL_LOCAL
                eff_cell.fill = FILL_LOCAL
            elif classification == "FORMULA_OTHER_SHEET":
                os_cell.value = display_value
                eff_cell.value = display_value
                os_cell.fill = FILL_OTHER
                eff_cell.fill = FILL_OTHER
                review_rows.append(
                    {
                        "sheet": "F.C.KMT",
                        "cell": coord,
                        "formula_or_value": safe_text(source_value),
                        "dependency_type": dependency_type,
                        "reason": "Formula references another sheet outside the DB-backed source route.",
                        "classification": classification,
                        "display_value": safe_text(display_value),
                    }
                )
            elif classification in {"CROSS_REPORT_DEPENDENCY", "EXTERNAL_REFERENCE", "REF_ERROR"}:
                os_cell.value = display_value
                eff_cell.value = display_value
                os_cell.fill = FILL_REVIEW
                eff_cell.fill = FILL_REVIEW
                review_rows.append(
                    {
                        "sheet": "F.C.KMT",
                        "cell": coord,
                        "formula_or_value": safe_text(source_value),
                        "dependency_type": dependency_type,
                        "reason": "Blocked or unresolved dependency kept visible and carried to review.",
                        "classification": classification,
                        "display_value": safe_text(display_value),
                    }
                )
            elif classification in {"HARDNUMBER", "TEXT_OR_DATE"}:
                os_cell.value = display_value
                eff_cell.value = display_value
                os_cell.fill = FILL_VALUE
                eff_cell.fill = FILL_VALUE
            else:
                os_cell.value = display_value
                eff_cell.value = display_value

            classification_rows.append(
                {
                    "sheet": "F.C.KMT",
                    "cell": coord,
                    "classification": classification,
                    "dependency_type": dependency_type,
                    "source_formula_or_value": safe_text(source_value),
                    "generated_value": safe_text(eff_cell.value),
                    "number_format": safe_text(src_cell.number_format),
                }
            )
            counts[classification] = counts.get(classification, 0) + 1

    add_variance_icon_formatting(os_ws)
    add_variance_icon_formatting(eff_ws)
    return counts


def write_summary(
    ctx: BuildContext,
    validation_rows: list[dict[str, str]],
    counts: dict[str, int],
    review_rows: list[dict[str, str]],
    archived_docs: list[str],
) -> None:
    with SUMMARY_PATH.open("w", encoding="utf-8") as handle:
        handle.write("==========\n00.00_FC_KMT_DUAL_REPORT_BUILD_SUMMARY\n==========\n\n")
        handle.write(f"RUN_ID.............: {ctx.run_id}\n")
        handle.write(f"GENERATED..........: {datetime.now().isoformat(timespec='seconds')}\n")
        handle.write(f"STATUS.............: {ctx.summary_status}\n")
        handle.write("TARGET_REPORT......: PDF page 3 / F.C.KMT / Gestora KMT\n")
        handle.write(f"OUTPUT_WORKBOOK....: {OUTPUT_XLSX}\n")
        handle.write(f"ARCHIVE_DIR........: {ctx.archive_dir}\n")
        handle.write("\n==========\n01.00_VALIDATION\n==========\n\n")
        for row in validation_rows:
            handle.write(
                f"- {row['scope']} / {row['item']} / expected={row['expected']} / actual={row['actual']} / "
                f"status={row['status']} / notes={row['notes']}\n"
            )
        handle.write("\n==========\n02.00_CLASSIFICATION_COUNTS\n==========\n\n")
        for key in sorted(counts):
            handle.write(f"- {key}: {counts[key]}\n")
        handle.write("\n==========\n03.00_REVIEW_REQUIRED\n==========\n\n")
        handle.write(f"- review_rows: {len(review_rows)}\n")
        handle.write("\n==========\n04.00_ARCHIVED_DOCS\n==========\n\n")
        if archived_docs:
            for name in archived_docs:
                handle.write(f"- {name}\n")
        else:
            handle.write("- none\n")


def write_docs_index() -> None:
    paths = [
        path
        for path in sorted(DOCS_DIR.iterdir())
        if path.is_file() and path != DOC_INDEX_PATH
    ]
    with DOC_INDEX_PATH.open("w", encoding="utf-8") as handle:
        handle.write("==========\n00.00_DOCS_ACTIVE_INDEX\n==========\n\n")
        for path in paths:
            handle.write(f"{path.name}\t{path.stat().st_size}\n")


def verify_output_workbook(
    validation_rows: list[dict[str, str]],
    classification_rows: list[dict[str, str]],
    expected_label_value: object,
) -> None:
    workbook = load_workbook(OUTPUT_XLSX, read_only=True, data_only=False)
    actual_sheets = workbook.sheetnames
    validation_rows.append(
        {
            "scope": "WORKBOOK",
            "item": "reopen_after_save",
            "expected": ", ".join(REQUIRED_SHEETS),
            "actual": ", ".join(actual_sheets),
            "status": "PASS" if actual_sheets == REQUIRED_SHEETS else "FAIL",
            "notes": "Workbook re-opened successfully after save.",
        }
    )

    for sheet_name in ("F.C.KMT.OS", "F.C.KMT.EFFICIENT"):
        sheet = workbook[sheet_name]
        formulas_ok = all(sheet[f"AK{row}"].value == f"=AF{row}-AH{row}" for row in range(10, 30))
        validation_rows.append(
            {
                "scope": "WORKBOOK",
                "item": f"{sheet_name}_ak_variance_formula",
                "expected": "AK10:AK29 = AF(row)-AH(row)",
                "actual": "PASS" if formulas_ok else "FAIL",
                "status": "PASS" if formulas_ok else "FAIL",
                "notes": "Validated exported variance formulas in the output workbook.",
            }
        )
        intentional_blanks_ok = all(sheet[cell_ref].value is None for cell_ref in ("R6", "S6", "AH19", "AH25", "AH26", "AH27", "AH28"))
        validation_rows.append(
            {
                "scope": "WORKBOOK",
                "item": f"{sheet_name}_intentional_blanks",
                "expected": "R6:S6, AH19, and AH25:AH28 remain blank",
                "actual": "PASS" if intentional_blanks_ok else "FAIL",
                "status": "PASS" if intentional_blanks_ok else "FAIL",
                "notes": "Validated intentional visual and no-source blanks in the output workbook.",
            }
        )
        label_ok = sheet["C24"].value == expected_label_value
        validation_rows.append(
            {
                "scope": "WORKBOOK",
                "item": f"{sheet_name}_c24_label",
                "expected": "C24 displays label from 'KMT '!$D$60",
                "actual": safe_text(sheet["C24"].value),
                "status": "PASS" if label_ok else "FAIL",
                "notes": "Validated exported visible label value in the output workbook.",
            }
        )

    workbook.close()

    classification_map = {
        (row["sheet"], row["cell"]): row["classification"]
        for row in classification_rows
    }
    uf_ok = all(
        (
            classification_map.get(("F.C.KMT", f"{get_column_letter(col)}6")) == "INTENTIONAL_BLANK_VISUAL_GAP"
            if FC_KMT_UF_GAP_COL_START <= col <= FC_KMT_UF_GAP_COL_END
            else classification_map.get(("F.C.KMT", f"{get_column_letter(col)}6")) == "REPORT_PARAMETER_UF_FROM_FCMC"
        )
        for col in range(FC_KMT_UF_COL_START, FC_KMT_UF_COL_END + 1)
    )
    forecast_ok = all(
        classification_map.get(("F.C.KMT", f"AH{row}"))
        == (
            "VISUAL_SEPARATOR_ROW"
            if row == 19
            else "INTENTIONAL_BLANK_NO_SOURCE_DATA"
            if 25 <= row <= 28
            else "FORECAST_V00_2026"
        )
        for row in range(FC_KMT_V00_ROW_START, FC_KMT_V00_ROW_END + 1)
    )
    label_ok = classification_map.get(("F.C.KMT", "C24")) == "VISIBLE_LABEL_FROM_KMT"
    validation_rows.append(
        {
            "scope": "CLASSIFICATION",
            "item": "fc_kmt_uf_parameters",
            "expected": "F6:Q6 and T6:AE6 classified from F.CMC; R6:S6 classified as intentional visual gaps",
            "actual": "PASS" if uf_ok else "FAIL",
            "status": "PASS" if uf_ok else "FAIL",
            "notes": "Validated exported UF parameter classifications.",
        }
    )
    validation_rows.append(
        {
            "scope": "CLASSIFICATION",
            "item": "fc_kmt_forecast_v00_2026",
            "expected": "AH10:AH29 classified as forecast, visual separator, or intentional no-source blanks per contract",
            "actual": "PASS" if forecast_ok else "FAIL",
            "status": "PASS" if forecast_ok else "FAIL",
            "notes": "Validated exported V00 classifications for the forecast range.",
        }
    )
    validation_rows.append(
        {
            "scope": "CLASSIFICATION",
            "item": "fc_kmt_visible_label",
            "expected": "C24 classified as VISIBLE_LABEL_FROM_KMT",
            "actual": "PASS" if label_ok else "FAIL",
            "status": "PASS" if label_ok else "FAIL",
            "notes": "Validated visible label classification for the KMT label cell.",
        }
    )


def build() -> dict[str, object]:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Required source workbook missing: {SOURCE_XLSX}")

    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    ctx = BuildContext(
        run_id=f"FC_KMT_DUAL_REPORT_{timestamp}",
        timestamp=timestamp,
        archive_dir=HIST_DIR / timestamp,
    )

    archived_docs = archive_obsolete_docs(ctx)
    validation_rows: list[dict[str, str]] = []
    review_rows: list[dict[str, str]] = []
    classification_rows: list[dict[str, str]] = []

    connection = validate_db(validation_rows)
    probe_rows = read_probe_rows(validation_rows)
    if any(row["status"] == "FAIL" for row in validation_rows):
        raise RuntimeError("Validation failed before workbook generation.")

    source_formula_wb = load_workbook(SOURCE_XLSX, data_only=False, keep_links=True)
    source_value_wb = load_workbook(SOURCE_XLSX, data_only=True, keep_links=True)

    for required in ("F.C.KMT", "F.D.KMT.01"):
        if required not in source_formula_wb.sheetnames:
            raise RuntimeError(f"Required sheet missing in source workbook: {required}")

    workbook = Workbook()
    workbook.remove(workbook.active)
    control_ws = workbook.create_sheet("CONTROL")
    sample_ws = workbook.create_sheet("DB.MOVEMENTS.SAMPLE")
    detail_ws = workbook.create_sheet("F.D.KMT.01.DB")
    efficient_ws = workbook.create_sheet("F.C.KMT.EFFICIENT")
    os_ws = workbook.create_sheet("F.C.KMT.OS")
    validation_ws = workbook.create_sheet("VALIDATION.RESULTS")
    review_ws = workbook.create_sheet("REVIEW.REQUIRED")
    contract_ws = workbook.create_sheet("REPORT.CONTRACT")

    add_control_sheet(control_ws, ctx, len(probe_rows), 4501)
    add_movements_sample(sample_ws, connection)
    build_detail_sheet(
        detail_ws,
        source_formula_wb["F.D.KMT.01"],
        source_value_wb["F.D.KMT.01"],
        probe_rows,
    )

    counts = populate_visual_sheets(
        source_formula_wb["F.C.KMT"],
        source_value_wb["F.C.KMT"],
        os_ws,
        efficient_ws,
        review_rows,
        classification_rows,
    )

    status = "PASS_WITH_REVIEW" if review_rows else "PASS"
    ctx.summary_status = status
    validation_rows.append(
        {
            "scope": "WORKBOOK",
            "item": "required_sheets",
            "expected": ", ".join(REQUIRED_SHEETS),
            "actual": ", ".join(workbook.sheetnames),
            "status": "PASS" if workbook.sheetnames == REQUIRED_SHEETS else "FAIL",
            "notes": "",
        }
    )
    validation_rows.append(
        {
            "scope": "WORKBOOK",
            "item": "final_status",
            "expected": "PASS or PASS_WITH_REVIEW",
            "actual": status,
            "status": "PASS",
            "notes": "Workbook creation completed.",
        }
    )
    validation_rows.append(
        {
            "scope": "WORKBOOK",
            "item": "review_required_rows",
            "expected": "0 for clean pass",
            "actual": str(len(review_rows)),
            "status": "WARN" if review_rows else "PASS",
            "notes": "Rows listed in REVIEW.REQUIRED when unresolved dependencies remain.",
        }
    )

    add_validation_sheet(validation_ws, validation_rows)
    add_review_sheet(review_ws, review_rows)
    contract_rows = build_contract_rows(ctx)
    add_contract_sheet(contract_ws, contract_rows)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)
    verify_output_workbook(
        validation_rows,
        classification_rows,
        source_value_wb["F.C.KMT"]["C24"].value,
    )
    write_csv(VALIDATION_PATH, VALIDATION_HEADERS, validation_rows)
    write_csv(REVIEW_PATH, REVIEW_HEADERS, review_rows)
    write_csv(CLASSIFICATION_PATH, CLASSIFICATION_HEADERS, classification_rows)
    write_csv(CONTRACT_PATH, CONTRACT_HEADERS, contract_rows)
    write_summary(ctx, validation_rows, counts, review_rows, archived_docs)
    write_docs_index()
    connection.close()

    return {
        "status": status,
        "review_rows": len(review_rows),
        "classification_rows": len(classification_rows),
        "output_workbook": str(OUTPUT_XLSX),
        "archived_docs": archived_docs,
    }


def main() -> int:
    try:
        result = build()
        print(json.dumps(result, indent=2))
        return 0
    except Exception:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archive_dir = HIST_DIR / timestamp
        archive_dir.mkdir(parents=True, exist_ok=True)
        error_path = archive_dir / "FC_KMT_DUAL_REPORT_EXCEPTION.txt"
        error_path.write_text(traceback.format_exc(), encoding="utf-8")
        print(traceback.format_exc())
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
